#!/usr/bin/env python
# -*- coding: utf-8 -*-

import ctypes
import os
import re
import sys
import threading
import tkinter as tk
from tkinter import filedialog, messagebox, ttk

import pandas as pd
import openpyxl

FONT      = 'Segoe UI'
FONT_SZ   = 13
FONT_SM   = 12
FONT_TITLE = 18
FONT_H1   = 14

def _init_dpi():
    if sys.platform != 'win32':
        return
    try:
        ctypes.windll.shcore.SetProcessDpiAwareness(2)
    except Exception:
        try:
            ctypes.windll.user32.SetProcessDPIAware()
        except Exception:
            pass

_init_dpi()
# Fluent Design Dark palette
FLUENT_BG_ROOT      = '#1b1b1b'
FLUENT_BG_SURFACE   = '#222222'
FLUENT_BG_ELEVATED  = '#2d2d2d'
FLUENT_BG_HOVER     = '#363636'
FLUENT_BG_ACTIVE    = '#3f3f3f'

FLUENT_TEXT         = '#f0f0f0'
FLUENT_TEXT_DIM     = '#aaaaaa'
FLUENT_TEXT_DISABLE = '#666666'

FLUENT_ACCENT       = '#60cdff'
FLUENT_ACCENT_HOVER = '#85d8ff'
FLUENT_ACCENT_PRESS = '#40b8e8'

FLUENT_BORDER       = '#383838'
FLUENT_BORDER_FOCUS = '#60cdff'
FLUENT_SELECTION    = '#154e66'
FLUENT_HEADER       = '#262626'
FLUENT_SCROLL       = '#4e4e4e'
FLUENT_SCROLL_HOVER = '#606060'
FLUENT_SEPARATOR    = '#333333'


def _apply_dark_style(root):
    style = ttk.Style(root)
    root.configure(bg=FLUENT_BG_ROOT)
    style.theme_use('clam')

    style.configure('.',
        background=FLUENT_BG_ROOT,
        foreground=FLUENT_TEXT,
        fieldbackground=FLUENT_BG_ELEVATED,
        bordercolor=FLUENT_BORDER,
        darkcolor=FLUENT_BG_ROOT,
        lightcolor=FLUENT_BG_ROOT,
        troughcolor=FLUENT_BG_SURFACE,
        selectbackground=FLUENT_ACCENT,
        selectforeground='#0d0d0d',
        arrowcolor=FLUENT_TEXT_DIM,
        font=(FONT, FONT_SZ),
    )

    style.configure('TFrame', background=FLUENT_BG_ROOT)
    style.configure('TLabelframe', background=FLUENT_BG_ROOT,
                    bordercolor=FLUENT_BORDER, relief='flat')
    style.configure('TLabelframe.Label', background=FLUENT_BG_ROOT,
                    foreground=FLUENT_TEXT, font=(FONT, FONT_SZ))
    style.configure('TLabel', background=FLUENT_BG_ROOT,
                    foreground=FLUENT_TEXT, font=(FONT, FONT_SZ))
    style.configure('Small.TLabel', font=(FONT, FONT_SM),
                    foreground=FLUENT_TEXT_DIM)
    style.configure('Bold.TLabel', font=(FONT, FONT_SZ, 'bold'))
    style.configure('H1.TLabel', font=(FONT, FONT_H1, 'bold'))
    style.configure('Title.TLabel', font=(FONT, FONT_TITLE, 'bold'))
    style.configure('Section.TLabel', font=(FONT, FONT_SZ + 1, 'bold'),
                    foreground=FLUENT_TEXT)

    style.configure('TButton',
        background=FLUENT_BG_ELEVATED,
        foreground=FLUENT_TEXT,
        bordercolor=FLUENT_BORDER,
        relief='flat',
        padding=(16, 7),
        font=(FONT, FONT_SZ),
        anchor='center',
    )
    style.map('TButton',
        background=[('active', FLUENT_BG_HOVER),
                    ('pressed', FLUENT_BG_ACTIVE),
                    ('disabled', FLUENT_BG_SURFACE)],
        foreground=[('active', FLUENT_TEXT),
                    ('disabled', FLUENT_TEXT_DISABLE)],
        bordercolor=[('active', FLUENT_BORDER_FOCUS),
                     ('focus', FLUENT_BORDER_FOCUS)],
    )

    style.configure('Accent.TButton',
        background=FLUENT_ACCENT,
        foreground='#0d0d0d',
        bordercolor=FLUENT_ACCENT,
        relief='flat',
        padding=(14, 7),
        font=(FONT, FONT_SZ, 'bold'),
    )
    style.map('Accent.TButton',
        background=[('active', FLUENT_ACCENT_HOVER),
                    ('pressed', FLUENT_ACCENT_PRESS),
                    ('disabled', FLUENT_BG_HOVER)],
        foreground=[('disabled', FLUENT_TEXT_DISABLE)],
    )

    style.configure('SegmentOn.TButton',
        background=FLUENT_ACCENT, foreground='#0d0d0d',
        bordercolor=FLUENT_ACCENT, relief='flat',
        padding=(10, 5), font=(FONT, FONT_SZ, 'bold'),
    )
    style.map('SegmentOn.TButton',
        background=[('active', FLUENT_ACCENT_HOVER),
                    ('pressed', FLUENT_ACCENT_PRESS)],
    )
    style.configure('SegmentOff.TButton',
        background=FLUENT_BG_ELEVATED, foreground=FLUENT_TEXT_DIM,
        bordercolor=FLUENT_BORDER, relief='flat',
        padding=(10, 5), font=(FONT, FONT_SZ),
    )
    style.map('SegmentOff.TButton',
        background=[('active', FLUENT_BG_HOVER),
                    ('pressed', FLUENT_BG_ACTIVE)],
        foreground=[('active', FLUENT_TEXT)],
    )

    style.configure('TEntry',
        fieldbackground=FLUENT_BG_ELEVATED,
        foreground=FLUENT_TEXT,
        insertcolor=FLUENT_TEXT,
        bordercolor=FLUENT_BORDER,
        relief='flat',
        padding=8,
        font=(FONT, FONT_SZ),
    )
    style.map('TEntry',
        fieldbackground=[('readonly', FLUENT_BG_SURFACE),
                         ('focus', FLUENT_BG_ELEVATED)],
        bordercolor=[('focus', FLUENT_BORDER_FOCUS)],
    )

    style.configure('TCombobox',
        fieldbackground=FLUENT_BG_ELEVATED,
        foreground=FLUENT_TEXT,
        arrowcolor=FLUENT_TEXT_DIM,
        bordercolor=FLUENT_BORDER,
        relief='flat',
        padding=6,
        font=(FONT, FONT_SZ),
    )
    style.map('TCombobox',
        fieldbackground=[('readonly', FLUENT_BG_ELEVATED),
                         ('focus', FLUENT_BG_ELEVATED)],
        bordercolor=[('focus', FLUENT_BORDER_FOCUS)],
    )
    root.option_add('*TCombobox*Listbox.background', FLUENT_BG_ELEVATED)
    root.option_add('*TCombobox*Listbox.foreground', FLUENT_TEXT)
    root.option_add('*TCombobox*Listbox.selectBackground', FLUENT_ACCENT)
    root.option_add('*TCombobox*Listbox.selectForeground', '#0d0d0d')
    root.option_add('*TCombobox*Listbox.font', (FONT, FONT_SZ))
    root.option_add('*TCombobox*Listbox.borderWidth', 0)
    root.option_add('*TCombobox*Listbox.relief', 'flat')

    style.configure('TCheckbutton',
        background=FLUENT_BG_ROOT,
        foreground=FLUENT_TEXT,
        font=(FONT, FONT_SZ),
    )
    style.map('TCheckbutton',
        background=[('active', FLUENT_BG_ROOT)],
        foreground=[('active', FLUENT_TEXT)],
    )

    style.configure('Vertical.TScrollbar',
        background=FLUENT_BG_SURFACE,
        troughcolor=FLUENT_BG_ROOT,
        bordercolor=FLUENT_BG_ROOT,
        arrowcolor=FLUENT_TEXT_DIM,
        gripcount=0,
        width=10,
        relief='flat',
    )
    style.map('Vertical.TScrollbar',
        background=[('active', FLUENT_SCROLL_HOVER)],
        arrowcolor=[('active', FLUENT_TEXT)],
    )
    style.configure('Horizontal.TScrollbar',
        background=FLUENT_BG_SURFACE,
        troughcolor=FLUENT_BG_ROOT,
        bordercolor=FLUENT_BG_ROOT,
        arrowcolor=FLUENT_TEXT_DIM,
        gripcount=0,
        width=10,
        relief='flat',
    )

    style.configure('Treeview',
        background=FLUENT_BG_SURFACE,
        foreground=FLUENT_TEXT,
        fieldbackground=FLUENT_BG_SURFACE,
        borderwidth=0,
        font=(FONT, 15),
        rowheight=48,
    )
    style.map('Treeview',
        background=[('selected', FLUENT_SELECTION)],
        foreground=[('selected', '#ffffff')],
    )
    style.configure('Treeview.Heading',
        background=FLUENT_HEADER,
        foreground=FLUENT_TEXT,
        relief='flat',
        font=(FONT, FONT_SZ, 'bold'),
        padding=(10, 6),
        borderwidth=0,
    )
    style.map('Treeview.Heading',
        background=[('active', FLUENT_BG_HOVER)],
    )

    style.configure('TProgressbar',
        background=FLUENT_ACCENT,
        troughcolor=FLUENT_BG_SURFACE,
        borderwidth=0,
        thickness=4,
    )

    style.configure('TSeparator', background=FLUENT_SEPARATOR)

    style.configure('TPanedwindow', background=FLUENT_BG_ROOT)
    style.configure('Sash', gripcount=0, sashthickness=2,
                    background=FLUENT_BORDER, bordercolor=FLUENT_BORDER)
    style.map('Sash', background=[('active', FLUENT_ACCENT)])

    style.configure('TNotebook', background=FLUENT_BG_ROOT, borderwidth=0)
    style.configure('TNotebook.Tab',
        background=FLUENT_BG_SURFACE,
        foreground=FLUENT_TEXT_DIM,
        padding=(18, 6),
        font=(FONT, FONT_SZ),
        borderwidth=0,
        relief='flat',
    )
    style.map('TNotebook.Tab',
        background=[('selected', FLUENT_BG_ROOT)],
        foreground=[('selected', FLUENT_ACCENT),
                    ('active', FLUENT_TEXT)],
        bordercolor=[('active', FLUENT_BORDER)],
    )

class PartsSearchApp:
    def __init__(self, root):
        self.root = root
        self.root.title('跨表零件号联合检索工具')
        self.root.minsize(1100, 650)
        self.root.state('zoomed')

        self.files = []
        self.file_cache = {}

        _apply_dark_style(root)
        self._build_ui()
        self.root.after(100, self._auto_import)

    def _build_ui(self):
        main = ttk.Frame(self.root)
        main.pack(fill=tk.BOTH, expand=True)

        self._build_top_bar(main)

        body = ttk.Frame(main)
        body.pack(fill=tk.BOTH, expand=True)
        body.grid_rowconfigure(0, weight=1)
        body.grid_columnconfigure(0, weight=30, uniform='main')
        body.grid_columnconfigure(2, weight=70, uniform='main')
        self.left_frame = ttk.Frame(body)
        self.left_frame.grid(row=0, column=0, sticky='ns')
        self._build_left_panel(self.left_frame)

        sep1 = ttk.Frame(body, width=1)
        sep1.configure(style='TSeparator')
        sep1.grid(row=0, column=1, sticky='ns')

        self.center_frame = ttk.Frame(body)
        self.center_frame.grid(row=0, column=2, sticky='nsew')
        self._build_center_panel(self.center_frame)

        self._build_status_bar(main)

        self._update_status()

    def _build_top_bar(self, parent):
        f = ttk.Frame(parent, padding=(16, 12, 16, 8))
        f.pack(fill=tk.X)
        ttk.Label(f, text='跨表零件号联合检索工具',
                  style='Title.TLabel').pack(side=tk.LEFT)
        ttk.Label(f, text='BOM 多文件联合查询',
                  foreground=FLUENT_TEXT_DIM, style='Small.TLabel').pack(
                      side=tk.LEFT, padx=(14, 0))
        ttk.Separator(parent, orient='horizontal').pack(fill=tk.X)

    def _build_left_panel(self, parent):
        c = ttk.Frame(parent, padding=(12, 8, 8, 8))
        c.pack(fill=tk.BOTH, expand=True)

        ttk.Label(c, text='基础设置', style='Section.TLabel').pack(
            anchor=tk.W, pady=(0, 6))

        ttk.Label(c, text='文件管理', style='Bold.TLabel').pack(anchor=tk.W, pady=(6, 4))
        btn_f = ttk.Frame(c)
        btn_f.pack(fill=tk.X)
        ttk.Button(btn_f, text='+ 导入文件', command=self._import_files,
                   style='Accent.TButton').pack(side=tk.LEFT)
        ttk.Button(btn_f, text='清除', command=self._clear_files).pack(side=tk.LEFT, padx=4)
        ttk.Button(btn_f, text='移除', command=self._remove_selected_files).pack(side=tk.LEFT)

        tree_f = ttk.Frame(c)
        tree_f.pack(fill=tk.BOTH, expand=True, pady=(4, 0))
        self.file_tree = ttk.Treeview(tree_f, columns=('name',), show='headings',
                                       height=3, selectmode='extended')
        self.file_tree.heading('name', text='文件名')
        self.file_tree.column('name', width=240)
        self.file_tree.tag_configure('alt_row', background=FLUENT_BG_SURFACE)
        fsb = ttk.Scrollbar(tree_f, orient=tk.VERTICAL, command=self.file_tree.yview)
        self.file_tree.configure(yscrollcommand=fsb.set)
        self.file_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        fsb.pack(side=tk.RIGHT, fill=tk.Y)

        ttk.Separator(c, orient='horizontal').pack(fill=tk.X, pady=(10, 6))

        ttk.Label(c, text='搜索设置', style='Bold.TLabel').pack(anchor=tk.W)

        ttk.Label(c, text='搜索列', padding=(0, 8, 0, 2)).pack(anchor=tk.W)
        col_row = ttk.Frame(c)
        col_row.pack(fill=tk.X)
        self.col_var = tk.StringVar(value='G')
        cb = ttk.Combobox(col_row, textvariable=self.col_var,
                          values=[chr(i) for i in range(65, 91)],
                          width=5, state='readonly')
        cb.pack(side=tk.LEFT)
        ttk.Label(col_row, text='A-第1列 G-第7列',
                  foreground=FLUENT_TEXT_DIM, style='Small.TLabel').pack(
                      side=tk.LEFT, padx=(6, 0))

        ttk.Label(c, text='零件号', padding=(0, 8, 0, 2)).pack(anchor=tk.W)
        self.part_entry = tk.Text(c, height=3, wrap=tk.WORD,
                                  bg=FLUENT_BG_ELEVATED, fg=FLUENT_TEXT,
                                  insertbackground=FLUENT_TEXT,
                                  font=(FONT, FONT_SZ),
                                  relief='flat', borderwidth=6,
                                  highlightthickness=1,
                                  highlightbackground=FLUENT_BORDER,
                                  highlightcolor=FLUENT_BORDER_FOCUS,
                                  padx=4, pady=4)
        self.part_entry.pack(fill=tk.X)

        self.ignore_case = tk.BooleanVar(value=True)
        ttk.Checkbutton(c, text='大小写不敏感',
                        variable=self.ignore_case).pack(anchor=tk.W)

        ttk.Label(c, text='检索模式', padding=(0, 8, 0, 2)).pack(anchor=tk.W)
        self.mode_var = tk.StringVar(value='多项检索')

        seg_frame = ttk.Frame(c)
        seg_frame.pack(fill=tk.X)
        modes = ['多项检索', '交集检索', '单一检索']
        self._mode_buttons = {}
        for m in modes:
            sty = 'SegmentOn.TButton' if m == self.mode_var.get() else 'SegmentOff.TButton'
            btn = ttk.Button(seg_frame, text=m, style=sty,
                             command=lambda v=m: self._set_mode(v))
            btn.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 1))
            self._mode_buttons[m] = btn

        self._mode_hint_var = tk.StringVar(
            value='逗号/分号/换行分隔多个零件号')
        ttk.Label(c, textvariable=self._mode_hint_var,
                  foreground=FLUENT_TEXT_DIM, style='Small.TLabel').pack(
                      anchor=tk.W, pady=(0, 4))

        ttk.Button(c, text='搜索', command=self._do_search,
                   style='Accent.TButton').pack(fill=tk.X, pady=(10, 4))

        ttk.Separator(c, orient='horizontal').pack(fill=tk.X, pady=(8, 6))

        ttk.Label(c, text='操作', style='Bold.TLabel').pack(anchor=tk.W)
        self.export_btn = ttk.Button(c, text='导出结果 (Excel)',
                                      command=self._export_results, state=tk.DISABLED)
        self.export_btn.pack(fill=tk.X, pady=(6, 0))
        ttk.Button(c, text='清除结果',
                    command=self._clear_results).pack(fill=tk.X, pady=(4, 0))

    def _build_center_panel(self, parent):
        c = ttk.Frame(parent, padding=(0, 0, 0, 0))
        c.pack(fill=tk.BOTH, expand=True)

        hdr = ttk.Frame(c, padding=(0, 4, 0, 4))
        hdr.pack(fill=tk.X)
        ttk.Label(hdr, text='查询结果', style='Section.TLabel').pack(side=tk.LEFT)
        self.result_count_label = ttk.Label(hdr, text='', foreground=FLUENT_TEXT_DIM,
                                            style='Small.TLabel')
        self.result_count_label.pack(side=tk.LEFT, padx=(10, 0))
        ttk.Label(hdr, text='双击查看完整表单',
                  foreground=FLUENT_TEXT_DIM, style='Small.TLabel').pack(side=tk.RIGHT)

        self.result_tabs = ttk.Notebook(c)
        self.result_tabs.bind('<<NotebookTabChanged>>', self._on_tab_changed)

        tab_all = ttk.Frame(self.result_tabs)
        self.result_tabs.add(tab_all, text='全部')
        self.result_tabs.pack(fill=tk.X, pady=(0, 4))

        tree_f = ttk.Frame(c)
        tree_f.pack(fill=tk.BOTH, expand=True)
        cols = ('零件号', '文件名', '工作表',
                '出现次数', '所在行', '匹配索引')
        self.result_tree = ttk.Treeview(tree_f, columns=cols, show='headings')
        col_w = [160, 220, 180, 90, 0, 0]
        for ci, w in zip(cols, col_w):
            self.result_tree.heading(ci, text=ci)
            if w:
                self.result_tree.column(ci, width=w, minwidth=60)
            else:
                self.result_tree.column(ci, width=0, stretch=False)
        self.result_tree.tag_configure('alt_row', background=FLUENT_BG_SURFACE)
        self.result_tree.bind('<Double-1>', self._on_result_double_click)
        self.result_tree.bind('<Delete>', lambda e: self._remove_selected_results())
        rsb = ttk.Scrollbar(tree_f, orient=tk.VERTICAL, command=self.result_tree.yview)
        self.result_tree.configure(yscrollcommand=rsb.set)
        self.result_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        rsb.pack(side=tk.RIGHT, fill=tk.Y)

        self._all_results = []

    def _auto_import(self):
        try:
            script_dir = os.path.dirname(os.path.abspath(__file__))
        except NameError:
            script_dir = os.getcwd()
        exts = ('.xlsx', '.xlsm', '.xls')
        for entry in os.listdir(script_dir):
            fp = os.path.join(script_dir, entry)
            if os.path.isfile(fp) and entry.lower().endswith(exts):
                if fp not in self.files:
                    self.files.append(fp)
                    self.file_tree.insert('', tk.END, values=(entry,))
        if self.files:
            self._stripe(self.file_tree)
            self._update_status()

    def _import_files(self):
        paths = filedialog.askopenfilenames(
            title='选择 Excel 文件',
            filetypes=[('Excel 文件', '*.xlsx *.xlsm *.xls'), ('All Files', '*.*')])
        if not paths:
            return
        added = 0
        for p in paths:
            if p not in self.files:
                self.files.append(p)
                self.file_tree.insert('', tk.END, values=(os.path.basename(p),))
                added += 1
        if added:
            self._stripe(self.file_tree)
        else:
            messagebox.showinfo('提示', '所选文件均已导入')
        self._update_status()

    def _remove_selected_files(self):
        selected = self.file_tree.selection()
        if not selected:
            return
        for item in reversed(selected):
            idx = self.file_tree.index(item)
            if idx < len(self.files):
                del self.files[idx]
            self.file_tree.delete(item)
        self.file_cache.clear()
        self._stripe(self.file_tree)
        self._update_status()

    def _clear_files(self):
        self.files.clear()
        self.file_cache.clear()
        for item in self.file_tree.get_children():
            self.file_tree.delete(item)
        self._update_status()

    def _clear_results(self):
        for item in self.result_tree.get_children():
            self.result_tree.delete(item)
        self._all_results = []
        for tab_id in self.result_tabs.tabs():
            self.result_tabs.forget(tab_id)
        tab_all = ttk.Frame(self.result_tabs)
        self.result_tabs.add(tab_all, text='全部')
        self.result_tabs.select(tab_all)
        self.result_count_label.configure(text='')
        self.export_btn.configure(state=tk.DISABLED)

    def _remove_selected_results(self):
        to_remove = set()
        for item in self.result_tree.selection():
            vals = self.result_tree.item(item)['values']
            to_remove.add(tuple(vals))
            self.result_tree.delete(item)

        self._all_results = [r for r in self._all_results
                             if tuple(r) not in to_remove]
        self._update_tab_counts()
        self._stripe(self.result_tree)

    def _stripe(self, tree):
        for i, item in enumerate(tree.get_children()):
            tree.item(item, tags=('alt_row',) if i % 2 == 0 else ())

    def _update_status(self):
        if not self.files:
            self.status_var.set('尚未导入文件 - 点击导入文件开始')
        else:
            self.status_var.set('已导入 {} 个文件'.format(len(self.files)))

    def _col_idx(self):
        return ord(self.col_var.get().upper()) - ord('A')

    @staticmethod
    def _norm(v):
        return '' if v is None else str(v).strip()

    def _match(self, cell_val, query):
        c = self._norm(cell_val)
        q = query
        if self.ignore_case.get():
            c = c.lower()
            q = q.lower()
        return c == q

    def _set_mode(self, mode):
        self.mode_var.set(mode)
        for m, btn in self._mode_buttons.items():
            if m == mode:
                btn.configure(style='SegmentOn.TButton')
            else:
                btn.configure(style='SegmentOff.TButton')
        hints = {
            '多项检索': '逗号/分号/换行分隔多个零件号',
            '交集检索': '输入多个零件号，仅显示同时存在的交集',
            '单一检索': '输入单个零件号精确检索',
        }
        self._mode_hint_var.set(hints.get(mode, hints['多项检索']))

    def _load_file(self, path):
        if path in self.file_cache:
            return self.file_cache[path]
        try:
            ext = os.path.splitext(path)[1].lower()
            if ext in ('.xlsx', '.xlsm'):
                sheets = pd.read_excel(path, sheet_name=None, dtype=str, engine='openpyxl')
            else:
                sheets = pd.read_excel(path, sheet_name=None, dtype=str, engine='xlrd')
        except Exception:
            try:
                sheets = {}
                wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
                for sn in wb.sheetnames:
                    ws = wb[sn]
                    data = []
                    for row in ws.iter_rows(values_only=True):
                        data.append([str(c) if c is not None else '' for c in row])
                    if data:
                        header = data[0]
                        body = data[1:]
                        ncols = max((len(r) for r in data), default=0)
                        for i, r in enumerate(body):
                            if len(r) < ncols:
                                body[i] = list(r) + [''] * (ncols - len(r))
                        if len(header) < ncols:
                            header = list(header) + ['Unnamed_{}'.format(j)
                                                     for j in range(len(header), ncols)]
                        sheets[sn] = pd.DataFrame(body, columns=header)
                    else:
                        sheets[sn] = pd.DataFrame()
                wb.close()
            except Exception:
                messagebox.showerror('错误',
                                     '无法读取文件: ' + os.path.basename(path))
                return None
        self.file_cache[path] = sheets
        return sheets

    def _do_search(self):
        raw = self.part_entry.get('1.0', 'end-1c').strip()
        if not raw:
            messagebox.showwarning('提示', '请输入要查询的零件号')
            return
        if not self.files:
            messagebox.showwarning('提示', '请先导入 Excel 文件')
            return

        mode = self.mode_var.get()
        queries = [q.strip() for q in re.split(r'[,;\n]+', raw) if q.strip()]
        if mode == '单一检索':
            queries = [raw.strip()]

        self._clear_results()
        col_idx = self._col_idx()
        col_letter = chr(ord('A') + col_idx)

        mode_labels = {'单一检索': '单一检索', '多项检索': '多项检索', '交集检索': '交集检索'}
        self.status_var.set('正在{} {} 个零件号 (列 {}) ...'.format(
            mode_labels.get(mode, ''), len(queries), col_letter))
        self.root.update_idletasks()

        def worker():
            if mode == '交集检索':
                results = self._search_intersect(queries, col_idx)
            else:
                results = []
                for fpath in self.files:
                    fname = os.path.basename(fpath)
                    sheets = self._load_file(fpath)
                    if sheets is None:
                        continue
                    for sname, df in sheets.items():
                        if df.empty or col_idx >= df.shape[1]:
                            continue
                        col = df.iloc[:, col_idx]
                        for query in queries:
                            mask = col.apply(lambda x: self._match(x, query))
                            matches = df[mask]
                            cnt = len(matches)
                            if cnt:
                                excel_rows = ','.join(str(i + 2) for i in matches.index)
                                df_indices = ','.join(str(i) for i in matches.index)
                                results.append((query, fname, sname, str(cnt),
                                                excel_rows, df_indices))
            self.root.after(0, lambda: self._show_results(results, queries, col_letter))
        threading.Thread(target=worker, daemon=True).start()

    def _search_intersect(self, queries, col_idx):
        results = []
        for fpath in self.files:
            fname = os.path.basename(fpath)
            sheets = self._load_file(fpath)
            if sheets is None:
                continue
            for sname, df in sheets.items():
                if df.empty or col_idx >= df.shape[1]:
                    continue
                col = df.iloc[:, col_idx]
                hits = {}
                all_matched = True
                for query in queries:
                    mask = col.apply(lambda x: self._match(x, query))
                    matches = df[mask]
                    if len(matches) == 0:
                        all_matched = False
                        break
                    excel_rows = ','.join(str(i + 2) for i in matches.index)
                    df_indices = ','.join(str(i) for i in matches.index)
                    hits[query] = (len(matches), excel_rows, df_indices)
                if all_matched:
                    for query in queries:
                        cnt, er, di = hits[query]
                        results.append((query, fname, sname, str(cnt), er, di))
        return results

    def _show_results(self, results, queries, col_letter):
        self._all_results = results

        self._rebuild_tabs(queries)

        if not results:
            self.status_var.set(
                '未找到匹配结果 - 搜索列 {}'.format(col_letter))
            self.result_count_label.configure(text='0 条记录')
            self.export_btn.configure(state=tk.DISABLED)
            return

        self._populate_tree(results)
        self._stripe(self.result_tree)

        found = set(r[0] for r in results)
        missing = [q for q in queries if q not in found]
        total = len(results)
        self.result_count_label.configure(text='共 {} 条记录'.format(total))
        status = '共找到 {} 条记录 (搜索列 {})'.format(total, col_letter)
        if missing:
            status += ' | 未找到: {}'.format(', '.join(missing))
        self.status_var.set(status)
        self.export_btn.configure(state=tk.NORMAL)

    def _rebuild_tabs(self, queries):
        for tab_id in self.result_tabs.tabs():
            self.result_tabs.forget(tab_id)

        tab_all = ttk.Frame(self.result_tabs)
        self.result_tabs.add(tab_all, text='全部 (0)')
        self.result_tabs.select(tab_all)

        unique_parts = sorted(set(q for q in queries))
        for p in unique_parts:
            tab_frame = ttk.Frame(self.result_tabs)
            self.result_tabs.add(tab_frame, text='{} (0)'.format(p[:20]))
        self._update_tab_counts()

    def _update_tab_counts(self):
        tabs = self.result_tabs.tabs()
        if not tabs:
            return
        total = len(self._all_results)
        self.result_tabs.tab(tabs[0], text='全部 ({})'.format(total))
        parts = {}
        for r in self._all_results:
            parts[r[0]] = parts.get(r[0], 0) + 1
        for i, tab_id in enumerate(tabs[1:], 1):
            text = self.result_tabs.tab(tab_id, 'text')
            pname = text.split(' (')[0] if ' (' in text else text
            cnt = parts.get(pname, 0)
            self.result_tabs.tab(tab_id, text='{} ({})'.format(pname[:20], cnt))

    def _on_tab_changed(self, event):
        sel = self.result_tabs.select()
        idx = self.result_tabs.index(sel)
        if idx == 0:
            self._populate_tree(self._all_results)
        else:
            tabs = self.result_tabs.tabs()
            tab_text = self.result_tabs.tab(tabs[idx], 'text')
            part_name = tab_text.split(' (')[0] if ' (' in tab_text else tab_text
            filtered = [r for r in self._all_results if r[0] == part_name]
            self._populate_tree(filtered)

    def _populate_tree(self, results):
        for item in self.result_tree.get_children():
            self.result_tree.delete(item)
        for r in results:
            self.result_tree.insert('', tk.END, values=r)

    def _on_result_double_click(self, event):
        sel = self.result_tree.selection()
        if not sel:
            return
        vals = self.result_tree.item(sel[0])['values']
        if len(vals) < 5:
            return
        part_num = str(vals[0])
        fname    = str(vals[1])
        sheet    = str(vals[2])
        excel_rows = str(vals[4]) if len(vals) > 4 else ''

        matched_file = next((f for f in self.files
                             if os.path.basename(f) == fname), None)
        if not matched_file:
            return

        first_row = excel_rows.split(',')[0].strip() if excel_rows else '1'
        self._open_excel_at_row(matched_file, sheet, first_row, part_num)

    def _open_excel_at_row(self, filepath, sheet_name, row_str, part_num):
        try:
            import pythoncom
            import win32com.client
            pythoncom.CoInitialize()
            excel = win32com.client.Dispatch("Excel.Application")
            excel.Visible = True
            wb = excel.Workbooks.Open(filepath)
            try:
                ws = wb.Sheets(sheet_name)
            except Exception:
                ws = wb.Sheets(1)
            ws.Activate()
            row = int(row_str) if row_str.isdigit() else 1
            scroll_row = max(1, row - 5)
            excel.ActiveWindow.ScrollRow = scroll_row
            ws.Cells(row, 1).Select()
            excel = None
            self.status_var.set(
                '已打开: {} -> {} 第{}行 (搜索: {})'.format(
                    os.path.basename(filepath), sheet_name, row, part_num))
        except ImportError:
            os.startfile(filepath)
            self.status_var.set(
                '已打开: {} -> {} 第{}行 (搜索: {})'.format(
                    os.path.basename(filepath), sheet_name, row_str, part_num))
        except Exception:
            os.startfile(filepath)
            self.status_var.set(
                '已打开: {} -> {} (搜索: {})'.format(
                    os.path.basename(filepath), sheet_name, part_num))

    def _build_status_bar(self, parent):
        btm = ttk.Frame(parent, padding=(16, 6, 16, 8))
        btm.pack(fill=tk.X, side=tk.BOTTOM)
        ttk.Separator(btm, orient='horizontal').pack(fill=tk.X, pady=(0, 4))
        self.status_var = tk.StringVar(value='尚未导入文件')
        ttk.Label(btm, textvariable=self.status_var,
                  foreground=FLUENT_TEXT_DIM, style='Small.TLabel').pack(side=tk.LEFT)

    def _export_results(self):
        if not self._all_results:
            return
        path = filedialog.asksaveasfilename(
            title='导出结果', defaultextension='.xlsx',
            filetypes=[('Excel 文件', '*.xlsx')])
        if not path:
            return
        try:
            rows = [r[:4] for r in self._all_results]
            df = pd.DataFrame(rows, columns=['零件号', '文件名',
                                             '工作表', '出现次数'])
            df.to_excel(path, index=False)
            messagebox.showinfo('提示',
                                '结果已导出到:\n' + os.path.basename(path))
        except Exception as e:
            messagebox.showerror('错误', '导出失败: ' + str(e))


def main():
    root = tk.Tk()
    try:
        root.iconbitmap(default='')
    except Exception:
        pass
    app = PartsSearchApp(root)
    root.mainloop()


if __name__ == '__main__':
    main()
