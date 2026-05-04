# main.py
"""
Copyright © 2025 Github:OUKUI All Rights Reserved.
"""
import tkinter as tk
from tkinter import font as tkfont
from tkinter import ttk
import datetime as dt
import os
import sys
import subprocess
# Optional: use Excel export (openpyxl). If not available, we skip exporting.
try:
    from openpyxl import Workbook
    OPENPYXL_AVAILABLE = True
except Exception:
    Workbook = None
    OPENPYXL_AVAILABLE = False
import re

# 打包为 exe 后用 sys.executable 定位 exe 所在目录，否则用脚本目录
if getattr(sys, 'frozen', False):
    BASE_DIR = os.path.dirname(sys.executable)
else:
    BASE_DIR = os.path.dirname(os.path.abspath(__file__))

# Log file for scan results
LOG_FILE = os.path.join(BASE_DIR, "scan.log")
# Model identifier for log file naming
MODEL = "ScannerApp"
# Directory to store CSV logs on exit
LOG_DIR = os.path.join(BASE_DIR, "logs")

class ScannerApp:
    def __init__(self, root):
        self.root = root
        self.root.title("现场扫码系统 - 单行显示版")
        
        # --- 窗口设置 ---
        self.root.state('zoomed')  # 最大化窗口
        
        # --- 深色主题配置 ---
        self.COLOR_BG_MAIN = "#1e1e1e"
        self.COLOR_BG_SEC = "#2d2d2d"
        self.COLOR_TEXT = "#ffffff"
        self.COLOR_TEXT_DIM = "#aaaaaa"
        self.COLOR_ACCENT = "#0078d7"
        # Counter display styling
        self.COLOR_COUNTER_BG = "#2c3e50"
        self.COLOR_COUNTER_FG = "#ecf0f1"
        self.font_counter = tkfont.Font(family="Microsoft YaHei", size=20, weight="bold")
        
        self.root.configure(bg=self.COLOR_BG_MAIN)
        # Ensure log file exists (simple trace)
        try:
            if not os.path.exists(LOG_FILE):
                with open(LOG_FILE, "w", encoding="utf-8") as f:
                    f.write(f"# Scan log initialized at {dt.datetime.now().isoformat()}\n")
        except Exception:
            pass

        # Init counters and session logs for real-time counting and CSV export
        self.total_count = 0
        self.ok_count = 0
        self.ng_count = 0
        self.session_logs = []  # list of dicts: time, target, scanned, result

        # --- 业务逻辑变量 ---
        self.target_text = ""
        self.ok_duration = 3
        self.ng_duration = 5
        
        # --- 字体 ---
        self.font_large = tkfont.Font(family="Microsoft YaHei", size=20, weight="bold")
        self.font_normal = tkfont.Font(family="Microsoft YaHei", size=14)
        self.font_kb = tkfont.Font(family="Arial", size=14, weight="bold")
        # 状态字体调大，适应单行显示
        self.font_status = tkfont.Font(family="Microsoft YaHei", size=80, weight="bold")

        self.create_login_screen()

        # Ensure window close handling to save logs
        self.root.protocol("WM_DELETE_WINDOW", self.on_close)
        self.original_target_text = ""
        self._touch_screen = self._has_touch_screen()

    # ================= 界面构建 =================

    def clear_screen(self):
        for widget in self.root.winfo_children():
            widget.destroy()

    def create_login_screen(self):
        """登录/设置界面"""
        self.clear_screen()
        
        # 标题
        header_frame = tk.Frame(self.root, bg=self.COLOR_BG_MAIN)
        header_frame.pack(fill='x', pady=10)
        
        tk.Label(header_frame, text="系统启动设置", font=self.font_large, 
                 bg=self.COLOR_BG_MAIN, fg=self.COLOR_TEXT).pack()
        tk.Label(header_frame, text="请输入目标编号及判定显示时间", 
                 font=self.font_normal, bg=self.COLOR_BG_MAIN, fg=self.COLOR_TEXT_DIM).pack(pady=5)

        # 目标编号输入
        input_frame = tk.Frame(self.root, bg=self.COLOR_BG_SEC, bd=1, relief="solid")
        input_frame.pack(pady=10, ipady=10, fill='x', padx=100)
        
        self.login_entry = tk.Entry(input_frame, font=("Arial", 20), justify='center',
                                    bg=self.COLOR_BG_SEC, fg=self.COLOR_TEXT, insertbackground='white',
                                    relief="flat")
        self.login_entry.pack(fill='x', padx=10, pady=5)
        tk.Button(self.root, text="⌨  显示键盘", font=("Microsoft YaHei", 13),
                  bg=self.COLOR_ACCENT, fg="white", relief="flat",
                  command=self.launch_windows_osk, padx=12, pady=4).pack(pady=(0, 6))

        # 时间设置区域
        time_frame = tk.Frame(self.root, bg=self.COLOR_BG_MAIN)
        time_frame.pack(fill='x', padx=100, pady=5)
        
        tk.Label(time_frame, text="OK显示时长(秒):", font=self.font_normal, 
                 bg=self.COLOR_BG_MAIN, fg=self.COLOR_TEXT).pack(side='left', padx=10)
        self.ok_time_entry = tk.Entry(time_frame, font=("Arial", 14), width=5, justify='center',
                                      bg=self.COLOR_BG_SEC, fg=self.COLOR_TEXT, relief="flat")
        self.ok_time_entry.insert(0, "3")
        self.ok_time_entry.pack(side='left', padx=5, ipady=5)

        tk.Label(time_frame, text="NG显示时长(秒):", font=self.font_normal,
                 bg=self.COLOR_BG_MAIN, fg=self.COLOR_TEXT).pack(side='left', padx=20)
        self.ng_time_entry = tk.Entry(time_frame, font=("Arial", 14), width=5, justify='center',
                                      bg=self.COLOR_BG_SEC, fg=self.COLOR_TEXT, relief="flat")
        self.ng_time_entry.insert(0, "5")
        self.ng_time_entry.pack(side='left', padx=5, ipady=5)
        # Start button for touchscreens
        start_btn = tk.Button(self.root, text="开始", font=("Microsoft YaHei", 22, "bold"),
                              bg="#27ae60", fg="white", command=self.start_app)
        start_btn.pack(pady=12)

        terms_frame = tk.Frame(self.root, bg=self.COLOR_BG_MAIN)
        terms_frame.pack(side='bottom', pady=10)
        tk.Label(terms_frame, text="若使用软件即表示您同意", font=("Microsoft YaHei", 10),
                 bg=self.COLOR_BG_MAIN, fg=self.COLOR_TEXT_DIM).pack(side='left')
        link = tk.Label(terms_frame, text="使用条款", font=("Microsoft YaHei", 10, "underline"),
                        bg=self.COLOR_BG_MAIN, fg=self.COLOR_ACCENT, cursor="hand2")
        link.pack(side='left')
        link.bind("<Button-1>", lambda e: LicenseWindow(self.root))
        tk.Label(terms_frame, text="。", font=("Microsoft YaHei", 10),
                 bg=self.COLOR_BG_MAIN, fg=self.COLOR_TEXT_DIM).pack(side='left')

    def start_app(self):
        """启动主程序"""
        raw = self.login_entry.get()
        val = raw.strip()
        self.original_target_text = raw.strip()
        if not val:
            self.login_entry.config(bg="#ffcccc")
            self.root.after(200, lambda: self.login_entry.config(bg=self.COLOR_BG_SEC))
            return
        
        self.target_text = val
        
        try:
            self.ok_duration = int(self.ok_time_entry.get())
            self.ng_duration = int(self.ng_time_entry.get())
        except ValueError:
            self.ok_duration = 2
            self.ng_duration = 4
            
        self.create_main_screen()

    def create_main_screen(self):
        """作业主界面"""
        self.clear_screen()
        
        # 顶部提示板
        self.top_label = tk.Label(self.root, text=f"目标编号：{self.original_target_text}", 
                                   font=("Microsoft YaHei", 28, "bold"), bg="#f1c40f", fg="black", pady=10)
        self.top_label.pack(fill='x')
        # 计数显示（总计/OK/NG）
        self.counter_label = tk.Label(self.root, text=self._counter_text(), font=self.font_counter,
                                      bg=self.COLOR_COUNTER_BG, fg=self.COLOR_COUNTER_FG)
        self.counter_label.pack(fill='x')

        # 中间区域
        main_frame = tk.Frame(self.root, bg=self.COLOR_BG_MAIN)
        main_frame.pack(fill='both', expand=True)

        tk.Label(main_frame, text="扫码输入区", font=self.font_normal, 
                 bg=self.COLOR_BG_MAIN, fg=self.COLOR_TEXT_DIM).pack(pady=10)
        
        self.scan_entry = tk.Entry(main_frame, font=("Arial", 30), justify='center',
                                   bg=self.COLOR_BG_SEC, fg=self.COLOR_TEXT, insertbackground='white',
                                   relief="flat")
        self.scan_entry.pack(pady=10, ipady=15, fill='x', padx=200)
        self.scan_entry.bind('<Return>', self.check_scan)
        
        # --- 修改：单一状态显示框 ---
        self.status_label = tk.Label(main_frame, text="● 待机中", font=self.font_status, 
                                     bg="#7f8c8d", fg="white", pady=80)
        self.status_label.pack(fill='both', expand=True, padx=50, pady=30)

        # 实时日志显示区
        log_frame = tk.Frame(main_frame, bg=self.COLOR_BG_MAIN)
        log_frame.pack(fill='both', expand=True, pady=10)
        self.log_text = tk.Text(log_frame, height=9, wrap='none', bg=self.COLOR_BG_SEC, fg=self.COLOR_TEXT,
                                font=("Consolas", 10), state='disabled')
        self.log_text.pack(side='left', fill='both', expand=True)
        log_scroll = tk.Scrollbar(log_frame, orient='vertical', command=self.log_text.yview)
        log_scroll.pack(side='right', fill='y')
        self.log_text.config(yscrollcommand=log_scroll.set)

        # 底部控制
        footer_frame = tk.Frame(self.root, bg=self.COLOR_BG_SEC, height=60)
        footer_frame.pack(fill='x', side='bottom')
        footer_frame.pack_propagate(False)
        
        tk.Button(footer_frame, text="重新设定", font=self.font_normal, 
                  command=self.create_login_screen,
                  bg="#c0392b", fg="white", relief="flat", padx=20, pady=5).pack(side='right', padx=20, pady=10)

        self.scan_entry.focus_set()

        # 取消自带虚拟按键盘实现，移除 create_keyboard 和 on_kb_click

    def check_scan(self, event=None):
        """扫码判定逻辑"""
        scanned_val = self.scan_entry.get().strip()
        
        if not scanned_val:
            return

        # 锁定输入框
        self.scan_entry.config(state='disabled')
        
        target_clean = str(self.target_text).strip()
        scanned_clean = str(scanned_val).strip()

        if scanned_clean == target_clean:
            self.set_status("OK", "#27ae60", self.ok_duration)
            self.log_scan(scanned_clean, "OK")
        else:
            self.set_status("NG", "#c0392b", self.ng_duration)
            self.log_scan(scanned_clean, "NG")

    def log_scan(self, scanned, result):
        """Log scan results to a simple log file (non-blocking)"""
        # Update counters and session log
        self.total_count += 1
        if result == "OK":
            self.ok_count += 1
        else:
            self.ng_count += 1
        self.session_logs.append({"time": dt.datetime.now().isoformat(),
                                  "target": self.original_target_text,
                                  "scanned": scanned,
                                  "result": result,
                                  "total": self.total_count,
                                  "ok": self.ok_count,
                                  "ng": self.ng_count})
        # Persist to log file
        try:
            with open(LOG_FILE, "a", encoding="utf-8") as f:
                f.write(f"{dt.datetime.now().isoformat()} | target={self.original_target_text} | scanned={scanned} | result={result} | total={self.total_count} | ok={self.ok_count} | ng={self.ng_count}\n")
        except Exception:
            pass
        # Update UI parts: log view and counters
        self._append_log_view(f"{dt.datetime.now().isoformat()} | target={self.target_text} | scanned={scanned} | result={result}")
        self._update_counters_view()

    def _has_touch_screen(self):
        try:
            result = subprocess.run(
                ["powershell", "-Command",
                 "(Get-WmiObject -Class Win32_PnPEntity | Where-Object {$_.Name -like '*touch*'}).Count -gt 0"],
                capture_output=True, text=True, timeout=3
            )
            return result.stdout.strip() == "True"
        except Exception:
            return False

    def launch_windows_osk(self):
        """触控屏用 TabTip，普通屏用 osk.exe（shell=True 确保系统能找到程序）。"""
        try:
            tabtip = r"C:\Program Files\Common Files\microsoft shared\ink\TabTip.exe"
            if os.path.exists(tabtip) and self._touch_screen:
                subprocess.Popen([tabtip], stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
            else:
                subprocess.Popen("osk.exe", shell=True)
        except Exception as e:
            print(f"无法打开键盘: {e}")

    def _safe_target(self):
        """Return a filesystem-safe representation of the current target_text."""
        t = self.original_target_text if getattr(self, 'original_target_text', '') else self.target_text
        # Replace any unsafe characters with underscore
        return re.sub(r'[^A-Za-z0-9_\-\.]', '_', t)

    def on_close(self):
        # Save logs to Excel before exiting
        self.save_logs_to_excel()
        # Try to close OSK if it's open
        try:
            subprocess.Popen(["taskkill", "/IM", "osk.exe", "/F"], stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
            subprocess.Popen(["taskkill", "/IM", "TabTip.exe", "/F"], stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
        except Exception:
            pass
        self.root.destroy()

    def save_logs_to_excel(self):
        if not self.session_logs:
            return
        if not OPENPYXL_AVAILABLE or Workbook is None:
            # Excel export unavailable; log and skip
            try:
                with open(LOG_FILE, "a", encoding="utf-8") as f:
                    f.write(f"{dt.datetime.now().isoformat()} | Excel export skipped: openpyxl not installed.\n")
            except Exception:
                pass
            return
        try:
            os.makedirs(LOG_DIR, exist_ok=True)
            safe_target = self._safe_target()
            filename = dt.datetime.now().strftime("%Y-%m-%d_%H-%M-%S") + f"_{safe_target}.xlsx"
            path = os.path.join(LOG_DIR, filename)
            wb = Workbook()
            ws = wb.active
            ws.title = "ScanLog"
            ws.append(["time", "target", "scanned", "result", "total", "ok", "ng"])
            for row in self.session_logs:
                ws.append([
                    row.get("time"),
                    row.get("target"),
                    row.get("scanned"),
                    row.get("result"),
                    row.get("total", ""),
                    row.get("ok", ""),
                    row.get("ng", ""),
                ])
            # Ensure target column is text to preserve leading zeros
            for r in range(2, ws.max_row+1):
                ws.cell(row=r, column=2).number_format = '@'
            wb.save(path)
        except Exception:
            pass

    def _append_log_view(self, line):
        if not hasattr(self, "log_text"):
            return
        self.log_text.config(state='normal')
        self.log_text.insert(tk.END, line + "\n")
        self.log_text.see(tk.END)
        self.log_text.config(state='disabled')

    def _counter_text(self):
        return f"Total: {self.total_count}  OK: {self.ok_count}  NG: {self.ng_count}"

    def _update_counters_view(self):
        if hasattr(self, "counter_label"):
            self.counter_label.config(text=self._counter_text())

    def set_status(self, text, color, duration):
        """单一文本框显示状态与倒计时（主线程 after 实现，tkinter 安全）"""
        self.status_label.config(text=f"{text} ({duration})", bg=color)
        self._countdown_tick(text, color, duration - 1)

    def _countdown_tick(self, text, color, remaining):
        if remaining > 0:
            self.status_label.config(text=f"{text} ({remaining})", bg=color)
            self.root.after(1000, self._countdown_tick, text, color, remaining - 1)
        else:
            # 倒计时结束：先解锁再清空，否则 disabled 状态下 delete 无效
            self.scan_entry.config(state='normal')
            self.scan_entry.delete(0, tk.END)
            self.scan_entry.focus_set()
            self.status_label.config(text="● 待机中", bg="#7f8c8d")

class LicenseWindow(tk.Toplevel):
    def __init__(self, parent):
        super().__init__(parent)

        self.title("许可协议")
        self.geometry("500x650")
        self.resizable(False, False)

        self.bg_color = "#1e1e1e"
        self.fg_color = "#e0e0e0"
        self.accent_color = "#3daee9"
        self.textbox_bg = "#2b2b2b"
        self.textbox_fg = "#cccccc"

        self.configure(bg=self.bg_color)

        self._create_header()
        self._create_terms()
        self._create_footer()

    def _create_header(self):
        header_frame = tk.Frame(self, bg=self.bg_color, pady=20, padx=20)
        header_frame.pack(fill="x")

        tk.Label(
            header_frame,
            text="版权声明",
            font=("Microsoft YaHei", 18, "bold"),
            bg=self.bg_color,
            fg=self.accent_color,
            anchor="w"
        ).pack(anchor="w")

        copyright_text = "Copyright © 2025 Github:OUKUI All Rights Reserved."
        tk.Label(
            header_frame,
            text=copyright_text,
            font=("Microsoft YaHei", 11),
            bg=self.bg_color,
            fg=self.fg_color,
            anchor="w"
        ).pack(anchor="w", pady=(10, 0))

    def _create_terms(self):
        text_container = tk.Frame(self, bg=self.bg_color, padx=20)
        text_container.pack(fill="both", expand=True)

        scrollbar = ttk.Scrollbar(text_container)
        scrollbar.pack(side="right", fill="y")

        self.terms_text = tk.Text(
            text_container,
            height=15,
            bg=self.textbox_bg,
            fg=self.textbox_fg,
            font=("Microsoft YaHei", 10),
            relief=tk.FLAT,
            padx=15,
            pady=15,
            yscrollcommand=scrollbar.set,
            wrap="word",
            state="disabled"
        )
        self.terms_text.pack(side="left", fill="both", expand=True)

        scrollbar.config(command=self.terms_text.yview)

        self._insert_terms_content()

    def _insert_terms_content(self):
        self.terms_text.config(state="normal")
        self.terms_text.delete("1.0", tk.END)

        terms_content = """《使用条款》
条款版本号：V2.0
最后更新日期：2026/5/3
著作权人：GitHub @OUKUI

软件使用条款

欢迎使用本软件。请您仔细阅读以下条款。您安装或使用本软件，即视为您已阅读、理解并同意受本条款的约束。如您不同意本条款的任何内容，请勿安装或使用本软件。

一、版权声明与禁止行为

本软件的全部著作权，包括但不限于源代码、程序架构、界面设计、运行逻辑、相关文档，以及由本软件运行所产生的数据记录、数据结构等，均归著作权人独立所有，受《中华人民共和国著作权法》及相关国际条约的严格保护。

未经著作权人书面授权，任何单位或个人不得实施以下行为：
（1）以任何方式复制、修改、改编、翻译本软件的全部或部分内容；
（2）通过互联网、移动存储、网络共享、预装、嵌入或其他任何渠道，对本软件进行分发、再分发、销售、出租、出借、许可、转让或向任何第三方提供；
（3）对本软件进行逆向工程、反编译、反汇编，或以其他任何方式试图获取或还原源代码、算法或架构；
（4）破解、规避、移除或绕过本软件中的任何技术保护措施、安全机制或身份验证系统；
（5）删除、修改、隐藏或移除本软件中的著作权声明、商标标识或其他权利信息；
（6）利用本软件的全部或部分内容制作衍生作品、插件或外挂，或将其整合到其他产品或服务中；
（7）超出著作权人明确书面授权的期限、设备数量、使用场景或功能范围使用本软件；
（8）协助、教唆或为他人实施上述任何行为提供便利。

二、授权使用限制

经著作权人书面授权的使用者，仅可在授权文件中载明的期限、设备数量、使用场景及功能范围内使用本软件，不得以转借、转租、转让、再授权、共享许可等方式使任何第三方直接或间接使用本软件，亦不得超出授权规模扩大使用。

著作权人有权根据实际情况，提前书面通知使用者变更或取消授权（通知可采取电子邮件、站内信等形式）。使用者应在通知载明的期限内停止使用本软件，并采取删除、销毁等措施。

三、违规使用后果

凡未获得授权擅自使用本软件，或授权被取消、到期后仍继续使用，以及实施本条款第一条所列任一禁止行为的，均构成侵犯软件著作权及相关权利的违法行为。

对上述违规行为，著作权人有权要求侵权方立即停止侵权，删除所有软件副本及相关数据记录，并按照侵权期间由本软件生成或处理的每一条数据记录0.5元的标准主张侵权赔偿；若该赔偿仍不足以弥补著作权人实际损失的，著作权人有权要求以实际损失为标准补足差额。

著作权人保留通过诉讼、仲裁等法律途径追究侵权方全部法律责任的权利，包括但不限于要求赔偿损失、承担著作权人为维权而支出的合理费用（含律师费、公证费、鉴定费等）。

四、其他条款

本条款的解释权归著作权人所有。若本条款中的任何内容与法律强制性规定相抵触，以法律规定为准，但其余条款的效力不受影响。

因使用本软件或本条款引起的任何争议，双方应友好协商解决；协商不成的，任何一方均有权向著作权人所在地有管辖权的人民法院提起诉讼。"""

        self.terms_text.insert(tk.END, terms_content)
        self.terms_text.config(state="disabled")

    def _create_footer(self):
        footer_frame = tk.Frame(self, bg=self.bg_color, pady=20)
        footer_frame.pack(fill="x")

        close_btn = tk.Button(
            footer_frame,
            text="我同意",
            font=("Microsoft YaHei", 11, "bold"),
            bg=self.accent_color,
            fg="white",
            activebackground="#2c8bb8",
            activeforeground="white",
            border=0,
            cursor="hand2",
            width=15,
            command=self.destroy
        )
        close_btn.pack()

if __name__ == "__main__":
    root = tk.Tk()
    app = ScannerApp(root)
    root.mainloop()
